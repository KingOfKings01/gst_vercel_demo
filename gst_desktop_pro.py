import time
import os
import sys
import openpyxl
import base64
import subprocess
from playwright.sync_api import sync_playwright

try:
    from captcha_solver import solve_captcha_from_image
except ImportError:
    # Manual stub if solver is missing
    def solve_captcha_from_image(img): return ""

from extract_gstin_data import extract_gstin_data

# ─────────────────────────────────────────────
# CONFIG & CREDIT SYSTEM
# ─────────────────────────────────────────────
MAX_CREDITS = 500
CREDIT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".credits")
SCRIPT_PATH = os.path.abspath(__file__)

def get_credits():
    if not os.path.exists(CREDIT_FILE):
        with open(CREDIT_FILE, 'w') as f: f.write(str(MAX_CREDITS))
        return MAX_CREDITS
    try:
        with open(CREDIT_FILE, 'r') as f: return int(f.read().strip())
    except:
        return 0

def use_credit():
    current = get_credits()
    if current > 0:
        with open(CREDIT_FILE, 'w') as f: f.write(str(current - 1))
        return current - 1
    return 0

def self_destruct():
    print("\n" + "!"*40)
    print("USAGE LIMIT REACHED (500 ROWS)")
    print("THIS SOFTWARE WILL NOW SELF-DESTRUCT.")
    print("!"*40 + "\n")
    time.sleep(3)
    
    # Simple self-deletion
    try:
        if os.path.exists(CREDIT_FILE): os.remove(CREDIT_FILE)
        # We can't easily delete ourselves while running on all OS perfectly, 
        # but we can try os.remove and then exit.
        if os.name == 'nt':
            # On Windows, we use a cmd trick to delete after exit
            subprocess.Popen(f'timeout /t 2 /nobreak > NUL && del "{SCRIPT_PATH}"', shell=True)
            sys.exit()
        else:
            os.remove(SCRIPT_PATH)
            sys.exit()
    except Exception as e:
        print(f"Self-destruct failed: {e}")
        sys.exit()

# ─────────────────────────────────────────────
# AUTOMATION LOGIC (PLAYWRIGHT)
# ─────────────────────────────────────────────
def run_automation(excel_path):
    credits = get_credits()
    if credits <= 0:
        self_destruct()
        return

    print(f"\n--- Desktop GST Pro | Credits Remaining: {credits} ---")
    abs_excel_path = os.path.abspath(excel_path)
    
    wb = openpyxl.load_workbook(abs_excel_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    
    try:
        gstin_col = headers.index("GSTIN") + 1
    except ValueError:
        print("Error: 'GSTIN' column missing in Excel file.")
        return

    col_map = {h: i+1 for i, h in enumerate(headers)}
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        
        for i in range(2, sheet.max_row + 1):
            if get_credits() <= 0:
                self_destruct()
                break
                
            gstin = str(sheet.cell(row=i, column=gstin_col).value).strip()
            if not gstin or gstin == "None": continue

            print(f"Row {i} | Processing {gstin}... (Rem: {get_credits()})")
            
            page.goto("https://services.gst.gov.in/services/searchtp")
            page.fill("#for_gstin", gstin)
            
            # Captcha Solve Loop
            is_solved = False
            for attempt in range(3):
                try:
                    captcha_img = page.wait_for_selector("#imgCaptcha", timeout=5000)
                    img_bytes = captcha_img.screenshot()
                    nparr = np.frombuffer(img_bytes, np.uint8)
                    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                    
                    captcha_text = solve_captcha_from_image(img)
                    if len(captcha_text) == 6:
                        page.fill("#fo-captcha", captcha_text)
                        page.click("#lotsearch")
                        
                        # Check for success
                        try:
                            page.wait_for_selector("#lottable", timeout=3000)
                            is_solved = True
                            break
                        except:
                            page.click("button[ng-click='refreshCaptcha()']")
                            time.sleep(1)
                except:
                    break
            
            if is_solved:
                html = page.content()
                gst_data = extract_gstin_data(html)
                
                for key, value in gst_data.items():
                    if key not in col_map:
                        col_map[key] = len(col_map) + 1
                        sheet.cell(row=1, column=col_map[key]).value = key
                    sheet.cell(row=i, column=col_map[key]).value = value
                
                wb.save(abs_excel_path)
                use_credit()
                time.sleep(1)
            else:
                print(f"Skipping {gstin} - Captcha Solve Failed.")

        browser.close()
        print("\nBatch Complete.")
        if get_credits() <= 0:
            self_destruct()

if __name__ == "__main__":
    # Pre-run check for cv2/numpy
    import cv2
    import numpy as np
    
    if len(sys.argv) < 2:
        print("Drag and drop your Excel file onto this script, or run: python gst_desktop_pro.py <excel_file>")
        time.sleep(5)
        sys.exit()
    
    run_automation(sys.argv[1])
