import time
import os
import openpyxl
import base64
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

try:
    from captcha_solver import solve_captcha_from_image
except ImportError:
    from .captcha_solver import solve_captcha_from_image

from extract_gstin_data import extract_gstin_data

# Ensure we use paths relative to the script location
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def get_path(relative_path):
    return os.path.join(SCRIPT_DIR, relative_path)

def init_driver(headless=True):
    """Initializes the Selenium WebDriver (Remote or Local)."""
    chrome_options = Options()
    if headless:
        chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")

    ws_endpoint = os.environ.get("BROWSER_WS_ENDPOINT")
    is_vercel = os.environ.get("VERCEL") == "1"

    if ws_endpoint:
        print(f"Connecting to remote browser at {ws_endpoint}...")
        # Note: Remote browser connection via Selenium often uses a different protocol/port
        # but many providers (like Browserless) support Selenium on the same endpoint.
        # Ensure the endpoint is a valid Selenium-compatible URL (usually http://.../wd/hub)
        return webdriver.Remote(command_executor=ws_endpoint, options=chrome_options)
    elif is_vercel:
        raise Exception(
            "BROWSER_WS_ENDPOINT is not set. Vercel deployments require a remote browser URL "
            "(e.g., from Browserless.io) to run Selenium."
        )
    else:
        print("Launching local Chrome...")
        # Local development assumes chromedriver is in PATH or managed by webdriver-manager
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            return webdriver.Chrome(service=service, options=chrome_options)
        except ImportError:
            # Fallback for systems with pre-installed chromedriver
            return webdriver.Chrome(options=chrome_options)

def solve_and_submit_captcha(driver, status_callback=None):
    """Wait for manual captcha solve (automated solve currently stubbed to save bundle size)."""
    # Pre-check: if results are already visible
    try:
        if driver.find_elements(By.ID, "lottable"):
            return True
    except:
        pass

    if status_callback: status_callback("Manual Captcha Solution Required - Waiting...")
    
    try:
        # Wait for the results table to appear after manual solve
        WebDriverWait(driver, 180).until(
            EC.presence_of_element_located((By.ID, "lottable"))
        )
        return True
    except Exception:
        return False

def run_batch_gst_search_excel(excel_path="gst_data.xlsx", status_callback=None):
    """Processes GSTINs from Excel using Selenium."""
    abs_excel_path = get_path(excel_path)
    if not os.path.exists(abs_excel_path):
        return

    wb = openpyxl.load_workbook(abs_excel_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    
    try:
        gstin_col = headers.index("GSTIN") + 1
        legal_name_col = headers.index("Legal Name of Business") + 1
    except ValueError:
        # Create columns if missing
        if "GSTIN" not in headers:
            if status_callback: status_callback("Error: 'GSTIN' column missing.")
            return
        legal_name_col = len(headers) + 1
        sheet.cell(row=1, column=legal_name_col).value = "Legal Name of Business"
        headers.append("Legal Name of Business")

    col_map = {h: i+1 for i, h in enumerate(headers)}
    driver = None

    try:
        driver = init_driver(headless=True)
        processed_count = 0
        
        for i in range(2, sheet.max_row + 1):
            gstin = str(sheet.cell(row=i, column=gstin_col).value).strip()
            legal_name = str(sheet.cell(row=i, column=legal_name_col).value).strip()
            
            if not gstin or gstin == "None" or gstin == "nan":
                continue
            if legal_name != "" and legal_name != "None" and legal_name != "nan":
                continue

            processed_count += 1
            if status_callback: status_callback(f"Processing {gstin}...")

            driver.get("https://services.gst.gov.in/services/searchtp")
            
            # Fill GSTIN
            gstin_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "for_gstin"))
            )
            gstin_input.clear()
            gstin_input.send_keys(gstin)
            
            if solve_and_submit_captcha(driver, status_callback=status_callback):
                html = driver.page_source
                gst_data = extract_gstin_data(html)
                
                for key, value in gst_data.items():
                    if key not in col_map:
                        col_map[key] = len(col_map) + 1
                        sheet.cell(row=1, column=col_map[key]).value = key
                    sheet.cell(row=i, column=col_map[key]).value = value
                
                wb.save(abs_excel_path)
                time.sleep(1)

        if status_callback: 
            status_callback("Task complete." if processed_count > 0 else "All rows already filled.")

    except Exception as e:
        print(f"!!! Selenium Error: {e}")
        if status_callback: status_callback(f"Error: {str(e)}")
    finally:
        if driver:
            driver.quit()
