import time
import os
import base64
import openpyxl
from playwright.sync_api import sync_playwright
from extract_gstin_data import extract_gstin_data

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def get_path(relative_path):
    return os.path.join(SCRIPT_DIR, relative_path)

# ─────────────────────────────────────────────
# CAPTCHA RELAY: Shared state between automation thread and Flask routes
# ─────────────────────────────────────────────
captcha_state = {
    "image_b64": None,      # Base64 captcha image for the web UI
    "solution": None,       # User's typed solution from the web UI
    "waiting": False,       # True when automation is waiting for user input
}

def get_captcha_state():
    return captcha_state

def submit_captcha_solution(solution):
    captcha_state["solution"] = solution

def init_browser(p):
    """Launches headless Chromium on the server."""
    print("Launching headless Chromium...")
    browser = p.chromium.launch(
        headless=True,
        args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"]
    )
    context = browser.new_context()
    page = context.new_page()
    return page, context, browser

def wait_for_captcha_solution(page, status_callback=None):
    """
    Screenshots the captcha, sends it to the web UI, and waits for the user to type a solution.
    """
    try:
        captcha_element = page.wait_for_selector("#imgCaptcha", timeout=10000)
    except:
        return None

    # Screenshot just the captcha image
    img_bytes = captcha_element.screenshot()
    captcha_state["image_b64"] = base64.b64encode(img_bytes).decode("utf-8")
    captcha_state["solution"] = None
    captcha_state["waiting"] = True

    if status_callback:
        status_callback("Captcha found! Please solve it in the web UI.")

    # Wait for user to submit solution (up to 3 minutes)
    timeout = 180
    start = time.time()
    while captcha_state["solution"] is None and (time.time() - start) < timeout:
        time.sleep(0.5)

    captcha_state["waiting"] = False
    solution = captcha_state["solution"]
    captcha_state["solution"] = None
    captcha_state["image_b64"] = None

    return solution

def run_batch_gst_search_excel(excel_path="gst_data.xlsx", status_callback=None):
    """Processes GSTINs from Excel with interactive captcha relay."""
    abs_excel_path = os.path.abspath(excel_path)
    if not os.path.exists(abs_excel_path):
        if status_callback: status_callback("File not found.")
        return

    wb = openpyxl.load_workbook(abs_excel_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    try:
        gstin_col = headers.index("GSTIN") + 1
    except ValueError:
        if status_callback: status_callback("Error: 'GSTIN' column missing.")
        return

    col_map = {h: i+1 for i, h in enumerate(headers)}

    with sync_playwright() as p:
        page, context, browser = init_browser(p)

        processed = 0
        total = sheet.max_row - 1
        for i in range(2, sheet.max_row + 1):
            gstin = str(sheet.cell(row=i, column=gstin_col).value).strip()
            if not gstin or gstin in ("None", "nan", ""): continue

            processed += 1
            if status_callback:
                status_callback(f"[{processed}/{total}] Opening search for {gstin}...")

            page.goto("https://services.gst.gov.in/services/searchtp")
            page.wait_for_load_state("networkidle")
            page.fill("#for_gstin", gstin)
            time.sleep(1)

            # Captcha relay loop
            is_solved = False
            for attempt in range(3):
                solution = wait_for_captcha_solution(page, status_callback)
                if not solution:
                    if status_callback: status_callback(f"No captcha solution received. Retrying...")
                    page.click("button[ng-click='refreshCaptcha()']")
                    time.sleep(1)
                    continue

                page.fill("#fo-captcha", solution)
                page.click("#lotsearch")

                # Check for error or success
                try:
                    page.wait_for_selector("#lottable", timeout=5000)
                    is_solved = True
                    break
                except:
                    if status_callback: status_callback(f"Wrong captcha. Retrying ({attempt+1}/3)...")
                    try:
                        page.click("button[ng-click='refreshCaptcha()']")
                    except:
                        pass
                    time.sleep(1)

            if is_solved:
                html = page.content()
                gst_data = extract_gstin_data(html)
                for key, value in gst_data.items():
                    if key not in col_map:
                        col_map[key] = len(col_map) + 1
                        sheet.cell(row=1, column=col_map[key]).value = key
                    sheet.cell(row=i, column=col_map[key]).value = value
                wb.save(abs_excel_path)
                if status_callback: status_callback(f"[{processed}/{total}] ✅ {gstin} - Data saved.")
            else:
                if status_callback: status_callback(f"[{processed}/{total}] ❌ {gstin} - Skipped.")

            time.sleep(1)

        browser.close()
        if status_callback: status_callback("🎉 All rows processed!")
