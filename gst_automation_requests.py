import requests
import time
import os
import openpyxl
from bs4 import BeautifulSoup
from captcha_solver import solve_captcha_from_image
import cv2
import numpy as np
from extract_gstin_data import extract_gstin_data

# Ensure we use paths relative to the script location
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def get_path(relative_path):
    return os.path.join(SCRIPT_DIR, relative_path)

def run_batch_gst_search_excel(excel_path="gst_data.xlsx", status_callback=None):
    """Processes GSTINs from Excel using pure HTTP requests (No Browser)."""
    abs_excel_path = get_path(excel_path)
    if not os.path.exists(abs_excel_path):
        return

    wb = openpyxl.load_workbook(abs_excel_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    
    try:
        gstin_col = headers.index("GSTIN") + 1
    except ValueError:
        if status_callback: status_callback("Error: 'GSTIN' column missing.")
        return

    col_map = {h: i+1 for i, h in enumerate(headers)}

    # Initialize Session
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://services.gst.gov.in/services/searchtp"
    })

    processed_count = 0
    for i in range(2, sheet.max_row + 1):
        gstin = str(sheet.cell(row=i, column=gstin_col).value).strip()
        if not gstin or gstin == "None": continue

        processed_count += 1
        if status_callback: status_callback(f"Requesting {gstin}...")

        try:
            # 1. Fetch search page to get cookies
            session.get("https://services.gst.gov.in/services/searchtp", timeout=15)
            
            # 2. Fetch Captcha
            captcha_url = "https://services.gst.gov.in/services/captcha?rnd=" + str(time.time())
            captcha_resp = session.get(captcha_url, timeout=15)
            
            # 3. Solve Captcha
            nparr = np.frombuffer(captcha_resp.content, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            captcha_text = solve_captcha_from_image(img)
            
            if len(captcha_text) != 6:
                if status_callback: status_callback(f"Retrying captcha for {gstin}...")
                continue

            # 4. POST search
            payload = {
                "gstin": gstin,
                "captcha": captcha_text
            }
            # Note: Portal logic might require session data or specific headers
            search_resp = session.post("https://services.gst.gov.in/services/searchtp", data=payload, timeout=20)
            
            html = search_resp.text
            gst_data = extract_gstin_data(html)
            
            if gst_data:
                for key, value in gst_data.items():
                    if key not in col_map:
                        col_map[key] = len(col_map) + 1
                        sheet.cell(row=1, column=col_map[key]).value = key
                    sheet.cell(row=i, column=col_map[key]).value = value
                
                wb.save(abs_excel_path)
            
            time.sleep(1)

        except Exception as e:
            if status_callback: status_callback(f"Error {gstin}: {str(e)}")
            continue

    if status_callback: 
        status_callback("Complete." if processed_count > 0 else "All filled.")
